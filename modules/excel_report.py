import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


# ==============================
# Excel Styles
# ==============================

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="4472C4"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

OK_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C"
)

CRITICAL_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)


def auto_fit(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = max_length + 2


def apply_common_style(ws):
    if ws.max_row < 1:
        return

    # Header style
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Border for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = ws.dimensions

    auto_fit(ws)


def apply_dashboard_status_style(ws):
    """
    DASHBOARD columns:
    1 #
    2 User
    3 Missing Hours
    4 Missing Days
    5 Logged Days
    6 Total Hours
    7 Required Hours/Day
    8 Working Days
    9 Expected Hours
    10 Missing %
    11 Status
    """

    status_col = 11

    if ws.max_row < 2:
        return

    for row in ws.iter_rows(min_row=2):
        status_cell = row[status_col - 1]
        status = status_cell.value

        if status == "OK":
            status_cell.fill = OK_FILL

        elif status == "Warning":
            status_cell.fill = WARNING_FILL

        elif status == "Critical":
            status_cell.fill = CRITICAL_FILL


def create_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)

    ws.append(headers)

    for row in rows:
        ws.append(row)

    apply_common_style(ws)

    if title == "DASHBOARD":
        apply_dashboard_status_style(ws)

    return ws


def write_info_sheet(
    wb,
    from_date,
    to_date,
    total_entries,
    redmine_count,
    min_hours
):
    ws = wb.active
    ws.title = "INFO"

    rows = [
        ["Item", "Value"],
        ["From Date", from_date],
        ["To Date", to_date],
        [
            "Generated At",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ],
        ["Redmine Count", redmine_count],
        ["Total Entries", total_entries],
        ["Default Min Hours Per Day", min_hours]
    ]

    for row in rows:
        ws.append(row)

    apply_common_style(ws)

    return ws


def generate_excel(
    result,
    output_folder,
    from_date,
    to_date,
    total_entries,
    redmine_count,
    min_hours
):
    os.makedirs(output_folder, exist_ok=True)

    report_time = datetime.now().strftime("%H%M%S")

    from_date_str = from_date.replace("-", "")
    to_date_str = to_date.replace("-", "")

    output_file = os.path.join(
        output_folder,
        f"timesheet_report_"
        f"{from_date_str}_"
        f"{to_date_str}_"
        f"{report_time}.xlsx"
    )

    wb = Workbook()

    write_info_sheet(
        wb,
        from_date,
        to_date,
        total_entries,
        redmine_count,
        min_hours
    )

    create_sheet(
        wb,
        "DASHBOARD",
        [
            "Rank",
            "User",
            "Missing Hours",
            "Missing Days",
            "Logged Days",
            "Total Hours",
            "Required Hours/Day",
            "Working Days",
            "Expected Hours",
            "Missing %",
            "Status"
        ],
        result["dashboard_rows"]
    )

    create_sheet(
        wb,
        "DETAIL",
        [
            "Date",
            "User",
            "Redmine",
            "Project",
            "Hours"
        ],
        result["detail_rows"]
    )

    create_sheet(
        wb,
        "UNDER_HOURS",
        [
            "Date",
            "User",
            "Total Hours",
            "Missing",
            "Required Hours"
        ],
        result["under_hours_rows"]
    )

    create_sheet(
        wb,
        "NO_LOG",
        [
            "Date",
            "User",
            "Redmine",
            "Required Hours"
        ],
        result["no_log_rows"]
    )

    create_sheet(
        wb,
        "SUMMARY",
        [
            "User",
            "Logged Days",
            "Total Hours",
            "Required Hours"
        ],
        result["summary_rows"]
    )

    create_sheet(
        wb,
        "MISSING_HOURS_SUMMARY",
        [
            "User",
            "Missing Days",
            "Missing Hours"
        ],
        result["missing_summary_rows"]
    )

    wb.save(output_file)

    return output_file