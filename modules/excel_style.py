from openpyxl.styles import (
    Font,
    Border,
    Side,
    PatternFill,
    Alignment
)


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="4472C4"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

OK_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C"
)

CRITICAL_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)


GROUP_FILL = PatternFill(
    fill_type="solid",
    fgColor="F2F2F2"
)


def apply_group_by_date_style(ws, date_col=1):
    if ws.max_row < 3:
        return

    current_date = None
    use_fill = False

    for row_idx in range(2, ws.max_row + 1):
        date_value = ws.cell(
            row=row_idx,
            column=date_col
        ).value

        if date_value != current_date:
            current_date = date_value
            use_fill = not use_fill

        if use_fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(
                    row=row_idx,
                    column=col_idx
                ).fill = GROUP_FILL

def auto_fit(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = max_length + 2


def apply_common_style(ws):
    if ws.max_row < 1:
        return

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    auto_fit(ws)


def apply_dashboard_status_style(ws):
    status_col = 11

    if ws.max_row < 2:
        return

    for row in ws.iter_rows(min_row=2):
        status_cell = row[status_col - 1]
        status = str(status_cell.value).strip()

        if status == "OK":
            status_cell.fill = OK_FILL

        elif status == "Warning":
            status_cell.fill = WARNING_FILL

        elif status == "Critical":
            status_cell.fill = CRITICAL_FILL

def apply_currency_format(ws, column_index):
    """
    Format number:
    10,000
    100,000
    1,000,000
    """

    if ws.max_row < 2:
        return

    for row in ws.iter_rows(min_row=2):
        cell = row[column_index - 1]

        if isinstance(
            cell.value,
            (int, float)
        ):
            cell.number_format = "#,##0"

def apply_penalty_status_style(ws):

    status_col = 5
    amount_col = 4

    apply_currency_format(
        ws,
        amount_col
    )

    if ws.max_row < 2:
        return

    for row in ws.iter_rows(min_row=2):

        status_cell = row[
            status_col - 1
        ]

        status = (
            str(status_cell.value)
            .strip()
            .upper()
        )

        if status in [
            "ĐÃ ĐÓNG",
            "DA DONG"
        ]:

            status_cell.fill = OK_FILL

        elif status in [
            "CHƯA ĐÓNG",
            "CHUA DONG"
        ]:

            status_cell.fill = CRITICAL_FILL

        elif status in [
            "MIỄN PHẠT",
            "MIEN PHAT"
        ]:

            status_cell.fill = WARNING_FILL