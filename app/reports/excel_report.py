import os
from datetime import datetime

from openpyxl import Workbook

from app.reports.excel_style import (
    apply_common_style,
    apply_dashboard_status_style,
    apply_penalty_status_style,
    apply_group_by_date_style
)

def create_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)

    ws.append(headers)

    for row in rows:
        ws.append(row)

    apply_common_style(ws)

    if title in [
        "DETAIL",
        "UNDER_HOURS",
        "NO_LOG",
        "PENALTY_REPORT"
    ]:
        apply_group_by_date_style(ws, date_col=1)

    if title == "DASHBOARD":
        apply_dashboard_status_style(ws)

    elif title == "PENALTY_REPORT":
        apply_penalty_status_style(ws)

    return ws


def write_info_sheet(
    wb,
    from_date,
    to_date,
    total_entries,
    redmine_count,
    min_hours
):
    ws = wb.active
    ws.title = "INFO"

    rows = [
        ["Item", "Value"],
        ["From Date", from_date],
        ["To Date", to_date],
        [
            "Generated At",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ],
        ["Redmine Count", redmine_count],
        ["Total Entries", total_entries],
        ["Default Min Hours Per Day", min_hours]
    ]

    for row in rows:
        ws.append(row)

    apply_common_style(ws)

    return ws


def generate_excel(
    result,
    output_folder,
    from_date,
    to_date,
    total_entries,
    redmine_count,
    min_hours
):
    os.makedirs(output_folder, exist_ok=True)

    report_time = datetime.now().strftime("%H%M%S")

    from_date_str = from_date.replace("-", "")
    to_date_str = to_date.replace("-", "")

    output_file = os.path.join(
        output_folder,
        f"timesheet_report_"
        f"{from_date_str}_"
        f"{to_date_str}_"
        f"{report_time}.xlsx"
    )

    wb = Workbook()

    write_info_sheet(
        wb,
        from_date,
        to_date,
        total_entries,
        redmine_count,
        min_hours
    )

    create_sheet(
        wb,
        "DASHBOARD",
        [
            "#",
            "User",
            "Missing Hours",
            "Missing Days",
            "Logged Days",
            "Leave Days",
            "Total Hours",
            "Required Hours",
            "Working Days",
            "Expected Hours",
            "Missing %",
            "Status"
        ],
        result["dashboard_rows"]
    )

    create_sheet(
        wb,
        "PENALTY_REPORT",
        [
            "Ngày",
            "Thành viên",
            "Vi phạm",
            "Tiền phạt",
            "Trạng thái",
            "Ghi nhận",
            "Đóng tiền",
            "Ghi chú"
        ],
        result["penalty_rows"]
    )

    create_sheet(
        wb,
        "DETAIL",
        [
            "Date",
            "User",
            "Redmine",
            "Project",
            "Issue ID",
            "Issue URL",
            "Hours"
        ],
        result["detail_rows"]
    )

    create_sheet(
        wb,
        "UNDER_HOURS",
        [
            "Date",
            "User",
            "Total Hours",
            "Missing",
            "Required Hours"
        ],
        result["under_hours_rows"]
    )

    create_sheet(
        wb,
        "OVER_HOURS",
        [
            "Date",
            "User",
            "Total Hours",
            "Required Hours",
            "Over Hours"
        ],
        result["over_hours_rows"]
    )

    create_sheet(
        wb,
        "NO_LOG",
        [
            "Date",
            "User",
            "Redmine",
            "Required Hours"
        ],
        result["no_log_rows"]
    )

    create_sheet(
        wb,
        "SUMMARY",
        [
            "User",
            "Logged Days",
            "Total Hours",
            "Required Hours"
        ],
        result["summary_rows"]
    )

    create_sheet(
        wb,
        "MISSING_HOURS_SUMMARY",
        [
            "User",
            "Missing Days",
            "Missing Hours"
        ],
        result["missing_summary_rows"]
    )

    wb.save(output_file)

    return output_file