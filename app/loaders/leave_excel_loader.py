import glob
import os
import re
from datetime import datetime, timedelta

from openpyxl import load_workbook


REQUIRED_HEADERS = [
    "Mã nhân viên",
    "Tên nhân viên",
    "Số ngày nghỉ",
    "Ngày bắt đầu nghỉ",
    "Ngày kết thúc nghỉ",
    "Người duyệt",
    "Ngày duyệt"
]


def normalize_text(value):
    return str(value).strip() if value is not None else ""


def parse_leave_date(value):
    """
    Input examples:
        15/06/2026(Cả ngày)
        10/06/2026(Buổi sáng)
        10/06/2026(Buổi chiều)
        15/06/2026
    Return:
        ("2026-06-15", "Cả ngày")
    """

    text = normalize_text(value)

    match = re.search(
        r"(\d{1,2}/\d{1,2}/\d{4})(?:\((.*?)\))?",
        text
    )

    if not match:
        return None, ""

    date_part = match.group(1)
    session = match.group(2) or ""

    date_obj = datetime.strptime(
        date_part,
        "%d/%m/%Y"
    )

    return (
        date_obj.strftime("%Y-%m-%d"),
        session.strip()
    )


def find_header_row(ws):
    for row_idx in range(1, ws.max_row + 1):
        values = [
            normalize_text(
                ws.cell(row=row_idx, column=col_idx).value
            )
            for col_idx in range(1, ws.max_column + 1)
        ]

        if all(header in values for header in REQUIRED_HEADERS):
            return row_idx, values

    return None, []


def build_header_index(headers):
    return {
        header: index + 1
        for index, header in enumerate(headers)
        if header
    }


def get_cell_value(ws, row_idx, header_index, header_name):
    col_idx = header_index.get(header_name)

    if not col_idx:
        return ""

    return ws.cell(row=row_idx, column=col_idx).value


def is_approved(ws, row_idx, header_index):
    approver = normalize_text(
        get_cell_value(
            ws,
            row_idx,
            header_index,
            "Người duyệt"
        )
    )

    approved_date = normalize_text(
        get_cell_value(
            ws,
            row_idx,
            header_index,
            "Ngày duyệt"
        )
    )

    return bool(approver and approved_date)


def get_standard_user(
    employee_code,
    employee_name,
    user_mapping
):
    code_key = normalize_text(
        employee_code
    ).lower()

    name_key = normalize_text(
        employee_name
    ).lower()

    standard_user = user_mapping.get(
        code_key
    )

    if standard_user:
        return standard_user

    return user_mapping.get(
        name_key
    )


def add_leave(
    user_leaves,
    date,
    user,
    hours
):
    if not date or not user or hours <= 0:
        return

    key = (
        date,
        user
    )

    user_leaves[key] = (
        user_leaves.get(key, 0)
        + hours
    )


def expand_leave_range(
    start_date,
    start_session,
    end_date,
    end_session,
    leave_days,
    working_hours_per_day,
    session_mapping
):
    """
    Return list of:
        [(date, leave_hours), ...]

    Handles common cases:
        - same day full-day / half-day
        - multiple full days
        - start/end sessions if present
    """

    result = []

    if not start_date:
        return result

    if not end_date:
        end_date = start_date

    start_obj = datetime.strptime(
        start_date,
        "%Y-%m-%d"
    )

    end_obj = datetime.strptime(
        end_date,
        "%Y-%m-%d"
    )

    if start_obj > end_obj:
        return result

    # Same day
    if start_date == end_date:
        session = start_session or end_session or "Cả ngày"

        hours = session_mapping.get(
            session,
            leave_days * working_hours_per_day
        )

        result.append(
            (
                start_date,
                hours
            )
        )

        return result

    current = start_obj

    while current <= end_obj:
        date_str = current.strftime("%Y-%m-%d")

        if current == start_obj:
            hours = session_mapping.get(
                start_session,
                working_hours_per_day
            )

        elif current == end_obj:
            hours = session_mapping.get(
                end_session,
                working_hours_per_day
            )

        else:
            hours = working_hours_per_day

        result.append(
            (
                date_str,
                hours
            )
        )

        current += timedelta(days=1)

    return result


def load_user_leaves_from_excel(
    leaves_config,
    user_mapping
):
    excel_config = leaves_config.get(
        "excel",
        {}
    )

    input_folder = excel_config.get(
        "input_folder",
        "resources/leave_exports"
    )

    file_pattern = excel_config.get(
        "file_pattern",
        "LeaveRequest*.xlsx"
    )

    working_hours_per_day = leaves_config.get(
        "working_hours_per_day",
        8
    )

    session_mapping = leaves_config.get(
        "session_mapping",
        {
            "Cả ngày": 8,
            "Buổi sáng": 4,
            "Buổi chiều": 4
        }
    )

    user_leaves = {}

    files = glob.glob(
        os.path.join(
            input_folder,
            file_pattern
        )
    )

    if not files:
        print(
            f"No leave excel files found: "
            f"{input_folder}/{file_pattern}"
        )
        return user_leaves

    for file_path in files:
        print(f"Reading leave excel: {file_path}")

        wb = load_workbook(
            file_path,
            data_only=True
        )

        for ws in wb.worksheets:
            header_row, headers = find_header_row(ws)

            if not header_row:
                continue

            header_index = build_header_index(headers)

            for row_idx in range(
                header_row + 1,
                ws.max_row + 1
            ):
                if not is_approved(
                    ws,
                    row_idx,
                    header_index
                ):
                    continue

                employee_code = get_cell_value(
                    ws,
                    row_idx,
                    header_index,
                    "Mã nhân viên"
                )

                employee_name = get_cell_value(
                    ws,
                    row_idx,
                    header_index,
                    "Tên nhân viên"
                )

                standard_user = get_standard_user(
                    employee_code,
                    employee_name,
                    user_mapping
                )

                if not standard_user:
                    continue

                leave_days_value = get_cell_value(
                    ws,
                    row_idx,
                    header_index,
                    "Số ngày nghỉ"
                )

                try:
                    leave_days = float(
                        str(leave_days_value)
                        .replace(",", ".")
                        .strip()
                    )
                except Exception:
                    leave_days = 1

                start_value = get_cell_value(
                    ws,
                    row_idx,
                    header_index,
                    "Ngày bắt đầu nghỉ"
                )

                end_value = get_cell_value(
                    ws,
                    row_idx,
                    header_index,
                    "Ngày kết thúc nghỉ"
                )

                start_date, start_session = parse_leave_date(
                    start_value
                )

                end_date, end_session = parse_leave_date(
                    end_value
                )

                leave_items = expand_leave_range(
                    start_date=start_date,
                    start_session=start_session,
                    end_date=end_date,
                    end_session=end_session,
                    leave_days=leave_days,
                    working_hours_per_day=working_hours_per_day,
                    session_mapping=session_mapping
                )

                for date, hours in leave_items:
                    add_leave(
                        user_leaves,
                        date,
                        standard_user,
                        hours
                    )

    return user_leaves